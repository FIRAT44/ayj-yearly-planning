import pandas as pd
import streamlit as st
from datetime import datetime
import matplotlib.pyplot as plt
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import tempfile
import os

def tab_tarihsel_analiz(st, conn):
    st.subheader("📈 Tarihsel Uçuş Süre Analizi")

    df = pd.read_sql_query("SELECT plan_tarihi, sure, gorev_tipi FROM ucus_planlari", conn, parse_dates=["plan_tarihi"])
    if df.empty:
        st.warning("Veri bulunamadı.")
        return

    df["sure_saat"] = pd.to_timedelta(df["sure"]).dt.total_seconds() / 3600

    # === TARİH ARALIĞI SEÇİMİ ===
    min_date = df["plan_tarihi"].min()
    max_date = df["plan_tarihi"].max()
    col1, col2 = st.columns(2)
    with col1:
        tarih1 = st.date_input("Başlangıç Tarihi", min_value=min_date, max_value=max_date, value=min_date)
    with col2:
        tarih2 = st.date_input("Bitiş Tarihi", min_value=min_date, max_value=max_date, value=max_date)

    tarih1 = pd.to_datetime(tarih1)
    tarih2 = pd.to_datetime(tarih2)
    # Tarih aralığına göre filtrele
    mask = (df["plan_tarihi"] >= tarih1) & (df["plan_tarihi"] <= tarih2)
    df_aralik = df[mask].copy()

    # ---- ZEKİ OTOMATİK İSTATİSTİKLER ----
    toplam_saat = df_aralik["sure_saat"].sum()
    ort_gunluk = df_aralik.groupby("plan_tarihi")["sure_saat"].sum().mean()
    min_gun = df_aralik.groupby("plan_tarihi")["sure_saat"].sum().min()
    max_gun = df_aralik.groupby("plan_tarihi")["sure_saat"].sum().max()
    hic_ucus_olmayan_gunler = pd.date_range(tarih1, tarih2).difference(df_aralik["plan_tarihi"].unique())
    populer_tip = df_aralik.groupby("gorev_tipi")["sure_saat"].sum().idxmax() if not df_aralik.empty else None

    st.success(f"""
        **Tarih Aralığı: {tarih1.date()} – {tarih2.date()}**
        \n- Toplam Uçuş Süresi: **{toplam_saat:.2f} saat**
        \n- Günlük Ortalama: **{ort_gunluk:.2f} saat**
        \n- En yüksek uçulan gün: **{max_gun:.2f} saat**
        \n- En düşük (uçuş yapılan) gün: **{min_gun:.2f} saat**
        \n- En çok uçulan görev tipi: **{populer_tip}**
        \n- Hiç uçuş olmayan gün(ler): {'Yok' if len(hic_ucus_olmayan_gunler)==0 else ', '.join(str(g.date()) for g in hic_ucus_olmayan_gunler[:5])}
        """)

    # ==== Tüm Görev Tipleri: Günlük Toplam ====
    st.markdown("## 🔷 Toplam Uçuş Süresi (Günlük)")
    df_gunluk_total = df_aralik.groupby("plan_tarihi")["sure_saat"].sum().reset_index().sort_values("plan_tarihi")

    st.line_chart(
        df_gunluk_total.set_index("plan_tarihi"),
        height=300,
        use_container_width=True
    )
    st.dataframe(df_gunluk_total, use_container_width=True)

    # Görev tipleri
    gorev_tipleri = df_aralik["gorev_tipi"].dropna().unique().tolist()
    gorev_tipleri.sort()

    selected_tips = st.multiselect(
        "Gösterilecek Görev Tiplerini Seçin", 
        options=gorev_tipleri, 
        default=gorev_tipleri,
        key="tarihsel_gorev_tipleri"
    )

    # ==== Her Görev Tipi: Günlük Grafik ====
    for tip in selected_tips:
        st.markdown(f"---\n## {tip}")
        df_tip = df_aralik[df_aralik["gorev_tipi"] == tip]
        df_gunluk = df_tip.groupby("plan_tarihi")["sure_saat"].sum().reset_index().sort_values("plan_tarihi")

        st.line_chart(
            df_gunluk.set_index("plan_tarihi"),
            height=250,
            use_container_width=True
        )
        st.dataframe(df_gunluk, use_container_width=True)
        # Akıllı özet
        st.info(
            f"- Toplam: {df_tip['sure_saat'].sum():.2f} saat | "
            f"Ortalama: {df_gunluk['sure_saat'].mean():.2f} saat/gün | "
            f"Gün sayısı: {df_gunluk['plan_tarihi'].nunique()}"
        )

    # ==== Excel Rapor Butonu (Tarih aralığına göre, grafikler + özet) ====
    st.markdown("### 📥 Excel Raporu (Grafik+Özet) İndir")
    excel_rapor_grafikli_indir(st, df_aralik, df_gunluk_total, selected_tips, tarih1, tarih2)
    tarihsel_akilli_ozet_panel(st, conn, tarih1, tarih2, df_aralik)

# -- GRAFİKLİ EXCEL EXPORT, AKILLI ÖZETLERLE --
def excel_rapor_grafikli_indir(st, df_aralik, df_gunluk_total, selected_tips, tarih1, tarih2):
    import matplotlib.pyplot as plt
    import tempfile
    import os
    import io
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage

    tmpdir = tempfile.mkdtemp()
    # GENEL TOPLAM GRAFİK
    path_total = os.path.join(tmpdir, "tarihsel_total.png")
    plt.figure(figsize=(8,3))
    plt.plot(df_gunluk_total["plan_tarihi"], df_gunluk_total["sure_saat"], marker="o")
    plt.title("Toplam Uçuş Süresi (Günlük)")
    plt.xlabel("Tarih")
    plt.ylabel("Saat")
    plt.tight_layout()
    plt.grid(True, linestyle="--", alpha=0.5)
    plt.savefig(path_total)
    plt.close()

    # Her görev tipi için grafikler
    tip_graf_paths = []
    for tip in selected_tips:
        df_tip = df_aralik[df_aralik["gorev_tipi"] == tip]
        df_gunluk = df_tip.groupby("plan_tarihi")["sure_saat"].sum().reset_index()
        if df_gunluk.empty:
            continue
        pth = os.path.join(tmpdir, f"tip_{tip}.png")
        plt.figure(figsize=(7,2.5))
        plt.plot(df_gunluk["plan_tarihi"], df_gunluk["sure_saat"], marker=".")
        plt.title(f"Görev Tipi: {tip}")
        plt.xlabel("Tarih")
        plt.ylabel("Saat")
        plt.grid(True, linestyle="--", alpha=0.5)
        plt.tight_layout()
        plt.savefig(pth)
        plt.close()
        tip_graf_paths.append((tip, pth))

    # EXCEL YAZ
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        # Özet sheet
        summary = {
            "Tarih Başlangıç": tarih1.date(),
            "Tarih Bitiş": tarih2.date(),
            "Toplam Uçuş Saati": df_aralik["sure_saat"].sum(),
            "Ortalama Günlük": df_aralik.groupby("plan_tarihi")["sure_saat"].sum().mean(),
            "En çok uçulan görev tipi": df_aralik.groupby("gorev_tipi")["sure_saat"].sum().idxmax() if not df_aralik.empty else "-",
        }
        pd.DataFrame([summary]).to_excel(writer, index=False, sheet_name="Özet")

        # Genel toplam/günlük sheet
        df_gunluk_total.to_excel(writer, index=False, sheet_name="Günlük Toplam")
        # Her görev tipi ayrı sheet
        for tip in selected_tips:
            df_tip = df_aralik[df_aralik["gorev_tipi"] == tip]
            df_gunluk = df_tip.groupby("plan_tarihi")["sure_saat"].sum().reset_index()
            df_gunluk.to_excel(writer, index=False, sheet_name=f"Tip_{tip[:25]}")

    # Grafik ekle
    excel_buffer.seek(0)
    wb = load_workbook(excel_buffer)
    ws = wb["Günlük Toplam"]
    img = XLImage(path_total)
    ws.add_image(img, "E2")

    for tip, pth in tip_graf_paths:
        ws_tip = wb[f"Tip_{tip[:25]}"]
        ws_tip.add_image(XLImage(pth), "E2")

    # Özet sheet'e de akıllı istatistikler eklenebilir, şimdilik tablo şeklinde bırakıldı
    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)

    st.download_button(
        label="📥 Seçilen Tarih Aralığına Göre Grafikli Excel Raporu İndir",
        data=out_buf,
        file_name=f"ucus_analiz_{tarih1.date()}_{tarih2.date()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



# === EKLE: 🎯 Akıllı Özet — Takvime Göre + Excel Raporu (HH:MM) ===
def tarihsel_akilli_ozet_panel(st, conn, tarih1, tarih2, df_aralik):
    """
    - Tüm metrikler SEÇİLEN TARİH ARALIĞINA göre hesaplanır.
    - 'Günlük (Uçuş Günleri)' YOK. Sadece 'Günlük (Takvim)' vardır.
    - Panelde görünen verilerin tamamı Excel raporu olarak indirilebilir.
    - Tüm saatler HH:MM formatında gösterilir ve rapora öyle yazılır (metin).
    """
    import pandas as pd
    from datetime import date
    import io

    def _to_hours(s):
        try:
            if pd.isna(s) or str(s).strip() == "" or str(s).strip().lower() == "nan":
                return 0.0
            td = pd.to_timedelta(str(s))
            return td.total_seconds() / 3600.0
        except Exception:
            try:
                return float(s)
            except Exception:
                return 0.0

    def _fmt_hhmm(hours_float: float) -> str:
        minutes = int(round(float(hours_float) * 60))
        hh, mm = divmod(minutes, 60)
        return f"{hh:02d}:{mm:02d}"

    # --- Güvenli giriş: df_aralik boşsa uyar ve çık ---
    if df_aralik is None or df_aralik.empty:
        st.warning("Seçilen tarih aralığında veri bulunamadı.")
        return

    # --- Seçilen aralık (tarih1–tarih2) verisi ---
    rng_min = pd.to_datetime(tarih1).normalize()
    rng_max = pd.to_datetime(tarih2).normalize()

    df_range = df_aralik.copy()
    df_range["plan_tarihi"] = pd.to_datetime(df_range["plan_tarihi"]).dt.normalize()
    if "sure_saat" not in df_range.columns and "sure" in df_range.columns:
        df_range["sure_saat"] = df_range["sure"].apply(_to_hours)

    # --- Günlük toplamlar (SEÇİLEN ARALIK) ---
    gunluk = (
        df_range.groupby("plan_tarihi")["sure_saat"]
        .sum()
        .sort_index()
    )
    total_hours    = gunluk.sum()
    flown_days     = gunluk.index.nunique()
    calendar_days  = (rng_max - rng_min).days + 1
    full_range     = pd.date_range(rng_min, rng_max, freq="D").normalize()
    not_flown_cnt  = max(len(full_range) - flown_days, 0)

    # --- Kalan plan & hedef (yalnız TAKVİM gününe bölünür) ---
    today       = pd.Timestamp(date.today()).normalize()
    start_point = max(today, rng_min)
    if start_point > rng_max:
        remaining_total = 0.0
        remaining_days  = 0
        kalan_tip = pd.DataFrame(columns=["gorev_tipi","kalan_saat"])
    else:
        kalan_df = df_range[df_range["plan_tarihi"] >= start_point]
        remaining_total = kalan_df["sure_saat"].sum()
        remaining_days  = (rng_max - start_point).days + 1

        kalan_tip = (
            kalan_df
            .assign(gorev_tipi=lambda d: d["gorev_tipi"].fillna("Bilinmiyor"))
            .groupby("gorev_tipi", dropna=False)["sure_saat"]
            .sum()
            .reset_index()
            .sort_values("sure_saat", ascending=False)
            .rename(columns={"sure_saat": "kalan_saat"})
        )

    hedef_gun_takvim = (remaining_total / remaining_days) if remaining_days > 0 else 0.0
    kalan_tip["gunluk_takvim"] = kalan_tip["kalan_saat"] / max(remaining_days, 1)

    # --- UI (tamamı seçilen aralığa göre) ---
    st.markdown(f"### 🎯 Akıllı Özet — Seçili Aralık: {rng_min.date()}–{rng_max.date()} (HH:MM)")

    c1, c2, c3 = st.columns(3)
    c1.metric("Toplam Plan (seçili aralık)", _fmt_hhmm(total_hours))
    c2.metric("Uçulan Gün (seçili aralık)", f"{flown_days} gün")
    c3.metric("Uçulmayan Gün (seçili aralık)", f"{not_flown_cnt} gün")

    st.info(
        f"**Kalan plan (başlangıç = {start_point.date()} → {rng_max.date()}): "
        f"{_fmt_hhmm(remaining_total)}**"
    )
    st.metric("Gerekli günlük (takvim gününe bölünmüş)", _fmt_hhmm(hedef_gun_takvim))

    # --- Görev tipi bazında yalnız 'Günlük (Takvim)' tablo ---
    st.markdown("### 🧮 Görev Tipi Bazında Günlük Gerekli (Takvim) — HH:MM")
    if not kalan_tip.empty:
        df_show = kalan_tip[["gorev_tipi","kalan_saat","gunluk_takvim"]].copy()
        df_show["kalan_saat"]    = df_show["kalan_saat"].apply(_fmt_hhmm)
        df_show["gunluk_takvim"] = df_show["gunluk_takvim"].apply(_fmt_hhmm)
        df_show.columns = ["Görev Tipi","Kalan Saat","Günlük (Takvim)"]
        st.dataframe(df_show, use_container_width=True)
    else:
        st.caption("Seçilen aralık için kalan görev tipi hedefi bulunmuyor.")

    # — Seçili aralığın kısa özeti (HH:MM)
    if not gunluk.empty:
        st.caption(
            f"Özet — Toplam: {_fmt_hhmm(total_hours)} | "
            f"Günlük ort. (uçulan): {_fmt_hhmm(gunluk.mean())} | "
            f"En yüksek gün: {_fmt_hhmm(gunluk.max())}"
        )

    # === 📥 RAPOR OLUŞTUR — Akıllı Özet (Excel) ===
    st.markdown("#### 📥 Akıllı Özet Raporu (Excel) — Seçilen Aralık")
    try:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            # 1) Akıllı Özet (tek satırlık metrikler)
            ozet_rows = [
                {"Alan": "Seçili Aralık", "Değer": f"{rng_min.date()} – {rng_max.date()}"},
                {"Alan": "Toplam Plan (seçili aralık)", "Değer": _fmt_hhmm(total_hours)},
                {"Alan": "Uçulan Gün", "Değer": f"{flown_days}"},
                {"Alan": "Uçulmayan Gün", "Değer": f"{not_flown_cnt}"},
                {"Alan": "Kalan Plan (başlangıç → bitiş)", "Değer": _fmt_hhmm(remaining_total)},
                {"Alan": "Gerekli Günlük (Takvim)", "Değer": _fmt_hhmm(hedef_gun_takvim)},
                {"Alan": "Özet (uçulan gün ort.)", "Değer": _fmt_hhmm(gunluk.mean() if len(gunluk)>0 else 0.0)},
                {"Alan": "Özet (en yüksek gün)", "Değer": _fmt_hhmm(gunluk.max() if len(gunluk)>0 else 0.0)},
            ]
            pd.DataFrame(ozet_rows).to_excel(writer, index=False, sheet_name="Akilli_Ozet")

            # 2) Tip Bazlı Günlük Gerekli (yalnız takvim)
            if not kalan_tip.empty:
                tip_df = kalan_tip[["gorev_tipi","kalan_saat","gunluk_takvim"]].copy()
                tip_df["kalan_saat"]    = tip_df["kalan_saat"].apply(_fmt_hhmm)
                tip_df["gunluk_takvim"] = tip_df["gunluk_takvim"].apply(_fmt_hhmm)
                tip_df.columns = ["Görev Tipi","Kalan Saat","Günlük (Takvim)"]
                tip_df.to_excel(writer, index=False, sheet_name="Tip_Bazli_Gunluk")

            # 3) Günlük Toplam (seçilen aralıktaki her gün)
            if not gunluk.empty:
                gt = gunluk.reset_index().rename(columns={"plan_tarihi": "Tarih", "sure_saat": "Toplam Saat"})
                gt["Toplam Saat"] = gt["Toplam Saat"].apply(_fmt_hhmm)
                gt.to_excel(writer, index=False, sheet_name="Gunluk_Toplam")

        buf.seek(0)
        st.download_button(
            label="📥 Akıllı Özet Raporunu İndir (Excel)",
            data=buf,
            file_name=f"akilli_ozet_{rng_min.date()}_{rng_max.date()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="akil_ozet_rapor_indir"
        )
    except Exception as e:
        st.error(f"Rapor oluşturulurken hata: {e}")

# --- Çağrı örneği (zaten mevcutsa dokunma) ---
# tab_tarihsel_analiz(...) fonksiyonunun en sonunda:
# tarihsel_akilli_ozet_panel(st, conn, tarih1, tarih2, df_aralik)
