import re, sqlite3
import pandas as pd
import numpy as np

# -------------------------------------------------------------------
# HIZLI: Yardımcılar
# -------------------------------------------------------------------


from datetime import datetime, date as _date

def _to_sql_ts(x, end_of_day=False) -> str:
    """
    x: date, datetime, pandas.Timestamp veya string olabilir.
    SQLite parametresi için 'YYYY-MM-DD HH:MM:SS' döner.
    end_of_day=True ise 23:59:59, değilse 00:00:00’a sabitler.
    """
    # pandas.Timestamp -> python datetime
    try:
        import pandas as pd  # mevcut zaten
        if isinstance(x, pd.Timestamp):
            x = x.to_pydatetime()
    except Exception:
        pass

    if isinstance(x, _date) and not isinstance(x, datetime):
        # sadece tarih ise başlangıç/bitiş saatini ekle
        return f"{x.isoformat()} {'23:59:59' if end_of_day else '00:00:00'}"

    if isinstance(x, datetime):
        if end_of_day:
            x = x.replace(hour=23, minute=59, second=59, microsecond=0)
        else:
            x = x.replace(microsecond=0)
        return x.strftime("%Y-%m-%d %H:%M:%S")

    # string gelirse normalize etmeye çalış
    try:
        dt = datetime.fromisoformat(str(x))
        if end_of_day:
            dt = dt.replace(hour=23, minute=59, second=59, microsecond=0)
        else:
            dt = dt.replace(microsecond=0)
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        # son çare: sadece tarih gibi kullan
        s = str(x).split()[0]
        return f"{s} {'23:59:59' if end_of_day else '00:00:00'}"



def _fmt_hhmmss(hours_float) -> str:
    try:
        sec = int(round(float(hours_float) * 3600))
    except Exception:
        sec = 0
    hh, rem = divmod(sec, 3600)
    mm, ss = divmod(rem, 60)
    return f"{hh:02d}:{mm:02d}:{ss:02d}"

def _fmt_hhmm_signed(hours_float) -> str:
    try:
        h = float(hours_float)
    except Exception:
        h = 0.0
    sign = "-" if h < 0 else ""
    h = abs(h)
    H = int(h)
    M = int(round((h - H) * 60))
    if M == 60:
        H += 1; M = 0
    return f"{sign}{H:02d}:{M:02d}"

def _format_time_cols(df, cols=None):
    df2 = df.copy()
    if cols is None:
        cols = [c for c in df2.columns
                if c.lower() in ["sure_saat", "toplam saat", "toplam", "kalan_saat", "gunluk_takvim", "y"]
                or c.startswith("A/C") or c in ["ME DUAL", "ME SIM", "SE SIM", "MCC SIM"]]
    for col in cols:
        if col in df2.columns:
            df2[col] = df2[col].apply(_fmt_hhmmss)
    return df2

def _ogr_kod_from_plan(s: str) -> str:
    s = str(s or "").strip()
    return s.split("-")[0].strip()

def _naeron_ogrenci_kodu_ayikla(pilot: str) -> str:
    """Naeron 'Öğrenci Pilot' → OZ/XXX kodu. (MCC dışı satırlarda kullanılır)"""
    if pd.isna(pilot): return ""
    pilot = str(pilot).strip()
    if pilot.startswith("OZ"):
        if pilot.count("-") >= 2:
            ikinci = [i for i, c in enumerate(pilot) if c == "-"][1]
            pilot = pilot[:ikinci].rstrip()
        return pilot
    else:
        return pilot.split("-")[0].strip()

# Block Time → saat (float)
def _to_hours_bt(s):
    if pd.isna(s) or s == "": return 0.0
    s = str(s).strip()
    # pandas to_timedelta ile hızlı parse dene
    td = pd.to_timedelta(s, errors="coerce")
    if pd.isna(td):
        # "1:30" gibi formata da izin
        parts = s.split(":")
        try:
            h = int(parts[0]); m = int(parts[1]) if len(parts) > 1 else 0
            s_ = int(parts[2]) if len(parts) > 2 else 0
            return h + m/60 + s_/3600
        except Exception:
            return 0.0
    return td.total_seconds()/3600.0

# -------------------------------------------------------------------
# CACHE'LENEN YÜKLEYİCİLER
# -------------------------------------------------------------------
def _create_indexes(conn: sqlite3.Connection):
    try:
        cur = conn.cursor()
        cur.execute("PRAGMA foreign_keys = ON;")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_plan_tarih ON ucus_planlari(plan_tarihi);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_plan_ogr ON ucus_planlari(ogrenci);")
        conn.commit()
    except Exception:
        pass

def _create_indexes_naeron(conn_naeron: sqlite3.Connection):
    try:
        cur = conn_naeron.cursor()
        # Köşeli/boşluklu isimler SQLite’ta kabul, ama index adı sade.
        cur.execute('CREATE INDEX IF NOT EXISTS idx_n_tarih2 ON naeron_ucuslar("Uçuş Tarihi 2");')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_n_gorev ON naeron_ucuslar("Görev");')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_n_pilot ON naeron_ucuslar("Öğrenci Pilot");')
        conn_naeron.commit()
    except Exception:
        pass

# Streamlit cache (yeniden başlatılsa da hızlı kalır)
try:
    import streamlit as st

    # ESKİ:
    # @st.cache_data(show_spinner=False)
    # def load_plan_slice(conn_path_or_conn, start, end):

    @st.cache_data(show_spinner=False)
    def load_plan_slice(_conn_path_or_conn, start, end):
        # start/end → ISO string
        start_s = _to_sql_ts(start, end_of_day=False)
        end_s   = _to_sql_ts(end,   end_of_day=True)

        # conn, path olabilir
        if isinstance(_conn_path_or_conn, sqlite3.Connection):
            conn = _conn_path_or_conn
            own = False
        else:
            conn = sqlite3.connect(_conn_path_or_conn)
            own = True
        try:
            _create_indexes(conn)
            q = """
            SELECT plan_tarihi, sure, gorev_tipi, ogrenci
            FROM ucus_planlari
            WHERE plan_tarihi >= ? AND plan_tarihi <= ?
            """
            dfp = pd.read_sql_query(q, conn, params=[start_s, end_s], parse_dates=["plan_tarihi"])
            return dfp
        finally:
            if own:
                conn.close()



    @st.cache_data(show_spinner=False)
    def load_naeron_slice(start, end):
        start_s = _to_sql_ts(start, end_of_day=False)
        end_s   = _to_sql_ts(end,   end_of_day=True)

        conn_n = sqlite3.connect("naeron_kayitlari.db")
        try:
            _create_indexes_naeron(conn_n)
            q = """
            SELECT
            "Uçuş Tarihi 2" AS ucus_tarihi,
            "Görev"         AS gorev,
            "Öğrenci Pilot" AS ogr_pilot,
            "Block Time"    AS block_time
            FROM naeron_ucuslar
            WHERE "Uçuş Tarihi 2" >= ? AND "Uçuş Tarihi 2" <= ?
            """
            dfn = pd.read_sql_query(q, conn_n, params=[start_s, end_s])
            return dfn
        finally:
            conn_n.close()


except Exception:
    # Streamlit yoksa cache’siz çalış
    def load_plan_slice(conn_path_or_conn, start, end):
        conn = conn_path_or_conn if isinstance(conn_path_or_conn, sqlite3.Connection) else sqlite3.connect(conn_path_or_conn)
        _create_indexes(conn)
        q = """
        SELECT plan_tarihi, sure, gorev_tipi, ogrenci
        FROM ucus_planlari
        WHERE plan_tarihi >= ? AND plan_tarihi <= ?
        """
        return pd.read_sql_query(q, conn, params=[start, end], parse_dates=["plan_tarihi"])

    def load_naeron_slice(start, end):
        conn_n = sqlite3.connect("naeron_kayitlari.db")
        _create_indexes_naeron(conn_n)
        q = """
        SELECT
          "Uçuş Tarihi 2" AS ucus_tarihi,
          "Görev"         AS gorev,
          "Öğrenci Pilot" AS ogr_pilot,
          "Block Time"    AS block_time
        FROM naeron_ucuslar
        WHERE "Uçuş Tarihi 2" >= ? AND "Uçuş Tarihi 2" <= ?
        """
        return pd.read_sql_query(q, conn_n, params=[start, end])

# -------------------------------------------------------------------
# HIZLI KARŞILAŞTIRMA ÇEKİRDEĞİ (tek geçiş)
# -------------------------------------------------------------------
def fast_plan_vs_actual(conn, tarih1, tarih2, only_planned_students=True):
    # 1) PLAN — sadece aralık
    dfp = load_plan_slice(conn, tarih1, tarih2)
    if dfp.empty:
        return (pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), set())

    dfp["sure_saat"] = pd.to_timedelta(dfp["sure"]).dt.total_seconds() / 3600.0
    dfp["gun"] = pd.to_datetime(dfp["plan_tarihi"]).dt.normalize()

    # Günlük toplam (PLAN)
    df_plan_daily = (dfp.groupby("gun")["sure_saat"].sum()
                       .reset_index()
                       .rename(columns={"sure_saat": "Planlanan"}))

    # 2) NAERON — tek seferde çek
    dfn = load_naeron_slice(tarih1, tarih2)
    if dfn.empty:
        # Uçuş yoksa sadece plan gösterilir
        df_cmp = df_plan_daily.copy()
        df_cmp["Gerçekleşen"] = 0.0
        df_cmp["Fark"] = df_cmp["Gerçekleşen"] - df_cmp["Planlanan"]
        return (df_cmp.rename(columns={"gun": "Tarih"}),
                df_plan_daily.rename(columns={"gun": "Tarih"}),
                pd.DataFrame(columns=["Ay","Planlanan","Gerçekleşen","Fark"]),
                set())

    # 2.a) Öğrenci filtresi
    plan_kod_set = set(dfp["ogrenci"].dropna().map(_ogr_kod_from_plan).unique())
    if not plan_kod_set:
        # plan içinde ogrenci kolonu boşsa tüm Naeron’u al
        only_planned_students = False

    # 2.b) MCC satırlarını çoklu öğrenciye patlat
    # 2.b) MCC satırlarını çoklu öğrenciye patlat  ✅ DÜZELTİLMİŞ
    dfn["is_mcc"] = dfn["gorev"].astype(str).str.upper().str.startswith("MCC")

    mcc = dfn[dfn["is_mcc"]].copy()
    if not mcc.empty:
        # Tüm öğrenci kodlarını çoklayıp orijinal satıra bağla
        extracted = (
            mcc["ogr_pilot"].astype(str).str.upper()
            .str.extractall(r"(\d{3}[A-Z]{2})")                # çoklu eşleşme
            .reset_index()                                     # -> ['level_0','match',0]
            .rename(columns={"level_0": "orig_idx", 0: "ogr_kod"})
        )
        if not extracted.empty:
            mcc_reset = mcc.reset_index().rename(columns={"index": "orig_idx"})
            # orig_idx üzerinden eşleştir
            mcc_exp = (
                mcc_reset.merge(extracted[["orig_idx", "ogr_kod"]], on="orig_idx", how="inner")
                .drop(columns=["orig_idx"])
            )
        else:
            mcc_exp = pd.DataFrame(columns=list(mcc.columns) + ["ogr_kod"])
    else:
        mcc_exp = pd.DataFrame(columns=list(dfn.columns) + ["ogr_kod"])

    # MCC dışı (tek öğrenci)
    other = dfn[~dfn["is_mcc"]].copy()
    other["ogr_kod"] = other["ogr_pilot"].map(_naeron_ogrenci_kodu_ayikla).fillna("")

    # Hepsini birleştir
    dfn_all = pd.concat([mcc_exp, other], ignore_index=True)

    if only_planned_students:
        dfn_all = dfn_all[dfn_all["ogr_kod"].isin(plan_kod_set)]

    # 2.c) tarih & saat
    dfn_all["ucus_tarihi"] = pd.to_datetime(dfn_all["ucus_tarihi"], errors="coerce")
    dfn_all = dfn_all.dropna(subset=["ucus_tarihi"])
    dfn_all["gun"] = dfn_all["ucus_tarihi"].dt.normalize()
    dfn_all["saat"] = dfn_all["block_time"].map(_to_hours_bt)

    # Günlük toplam (GERÇEK)
    df_act_daily = (dfn_all.groupby("gun")["saat"].sum()
                      .reset_index()
                      .rename(columns={"saat": "Gerçekleşen"}))

    # 3) BİRLEŞİK — dış birleştirme (plan var/ uçuş var)
    df_cmp = (pd.merge(df_plan_daily, df_act_daily, on="gun", how="outer")
                .fillna(0.0)
                .sort_values("gun"))
    df_cmp["Fark"] = df_cmp["Gerçekleşen"] - df_cmp["Planlanan"]

    # 4) Aylık özet
    tmp = df_cmp.copy()
    tmp["Ay"] = tmp["gun"].dt.to_period("M").dt.to_timestamp()
    aylik = (tmp.groupby("Ay")[["Planlanan","Gerçekleşen"]].sum().reset_index())
    aylik["Fark"] = aylik["Gerçekleşen"] - aylik["Planlanan"]

    return (df_cmp.rename(columns={"gun":"Tarih"}),
            df_plan_daily.rename(columns={"gun":"Tarih"}),
            aylik, plan_kod_set)

# -------------------------------------------------------------------
# ANA TAB (Hızlı)
# -------------------------------------------------------------------
def tab_ihtiyac_analizi_karsilastirma(st, conn):

    st.subheader("⚡ Tarihsel Uçuş Süre Analizi (Hızlandırılmış)")


    # Tarih aralığı — sadece plan tablosundan min/max al
    df_all = pd.read_sql_query("SELECT MIN(plan_tarihi) AS mn, MAX(plan_tarihi) AS mx FROM ucus_planlari", conn,
                               parse_dates=["mn","mx"])
    if df_all.empty or pd.isna(df_all.iloc[0]["mn"]) or pd.isna(df_all.iloc[0]["mx"]):
        st.warning("Plan verisi bulunamadı."); return
    min_date = df_all.iloc[0]["mn"]; max_date = df_all.iloc[0]["mx"]

    c1, c2 = st.columns(2)
    with c1:
        tarih1 = st.date_input("Başlangıç", min_value=min_date, max_value=max_date, value=min_date, key="fast_t1")
    with c2:
        tarih2 = st.date_input("Bitiş", min_value=min_date, max_value=max_date, value=max_date, key="fast_t2")
    tarih1 = pd.to_datetime(tarih1); tarih2 = pd.to_datetime(tarih2)

    # Filtre: sadece planda görünen öğrencilerle mi gerçekleşeni sayalım?
    only_planned = st.toggle("Gerçekleşeni sadece bu aralıkta planı bulunan öğrencilerle sınırla (önerilir)", True)

    # Hızlı çekirdek
    df_cmp, df_plan_daily, aylik, plan_kod_set = fast_plan_vs_actual(conn, tarih1, tarih2, only_planned_students=only_planned)

    if df_cmp.empty:
        st.info("Seçili aralıkta gösterilecek veri bulunamadı."); return

    # Akıllı özet (hızlı)
    st.markdown("### 🧠 Özet")
    tot_plan = df_plan_daily["Planlanan"].sum()
    tot_act  = df_cmp["Gerçekleşen"].sum()
    c1, c2, c3 = st.columns(3)
    c1.metric("Toplam Plan", _fmt_hhmmss(tot_plan))
    c2.metric("Toplam Gerçekleşen", _fmt_hhmmss(tot_act))
    c3.metric("Fark", _fmt_hhmm_signed(tot_act - tot_plan))

    # Günlük çizgi grafiği
    st.markdown("## 📈 Günlük — Planlanan vs Gerçekleşen")
    st.line_chart(
        df_cmp.set_index("Tarih")[["Planlanan","Gerçekleşen"]],
        height=320, use_container_width=True
    )

    # Günlük tablo (metin format)
    show = df_cmp.copy()
    show["Planlanan"]   = show["Planlanan"].apply(_fmt_hhmmss)
    show["Gerçekleşen"] = show["Gerçekleşen"].apply(_fmt_hhmmss)
    show["Fark"]        = show["Fark"].apply(_fmt_hhmm_signed)
    st.dataframe(show, use_container_width=True)

    # Aylık bar + tablo
    st.markdown("## 🗓️ Aylık Özet — Planlanan vs Gerçekleşen")
    st.bar_chart(aylik.set_index("Ay")[["Planlanan","Gerçekleşen"]], height=320, use_container_width=True)
    aylik_show = aylik.copy()
    aylik_show["Planlanan"]   = aylik_show["Planlanan"].apply(_fmt_hhmmss)
    aylik_show["Gerçekleşen"] = aylik_show["Gerçekleşen"].apply(_fmt_hhmmss)
    aylik_show["Fark"]        = aylik_show["Fark"].apply(_fmt_hhmm_signed)
    st.dataframe(aylik_show, use_container_width=True)

    # İsteğe bağlı: Plan breakdown (hafif)
    with st.expander("➕ Plan kırılımı (A/C vs SIM ve 4 görev) — (uygulamada plan üzerinden hızlı hesap)"):
        # Sadece plan verisinden, önceki yavaş kısımları yapısal olarak hafifletip gösteriyoruz
        # (df_plan_daily zaten var; şimdi sadece kategori bazlı planı çıkaracağız)
        dfp = load_plan_slice(conn, tarih1, tarih2).copy()
        if not dfp.empty:
            dfp["sure_saat"] = pd.to_timedelta(dfp["sure"]).dt.total_seconds()/3600.0
            dfp["ay"] = pd.to_datetime(dfp["plan_tarihi"]).dt.to_period("M").dt.to_timestamp()

            def _norm(s: str) -> str:
                return re.sub(r"\s+", " ", str(s).strip().upper())
            DA20_KEYS   = {"SE DUAL DA", "SE PIC"}
            SONACA_KEYS = {"SE DUAL SONACA"}
            AUPRT_KEYS  = {"AUPRT"}

            def map_kategori(gorev_tipi: str) -> str:
                s = _norm(gorev_tipi)
                if s == "ME DUAL": return "ME DUAL"
                if s.startswith("ME SIM"):  return "ME SIM"
                if s.startswith("SE SIM"):  return "SE SIM"
                if s.startswith("MCC SIM"): return "MCC SIM"
                if s in DA20_KEYS:    return "A/C – DA20"
                if s in SONACA_KEYS:  return "A/C – SONACA"
                if s in AUPRT_KEYS:   return "A/C – AUPRT"
                return "A/C – DİĞER"

            dfp["kategori_detay"] = dfp["gorev_tipi"].map(map_kategori)
            SIM_SET = {"ME SIM", "SE SIM", "MCC SIM"}

            # A/C vs SIM — aylık (Plan)
            aylik_ac  = (dfp[~dfp["kategori_detay"].isin(SIM_SET)].groupby("ay")["sure_saat"].sum())
            aylik_sim = (dfp[dfp["kategori_detay"].isin(SIM_SET)].groupby("ay")["sure_saat"].sum())
            piv_ac_sim = pd.concat([aylik_ac.rename("A/C"), aylik_sim.rename("SIM")], axis=1).fillna(0.0).sort_index()
            st.bar_chart(piv_ac_sim, height=280, use_container_width=True)
            st.dataframe(_format_time_cols(piv_ac_sim.reset_index().rename(columns={"ay":"Ay"})), use_container_width=True)

            # 4 görev — aylık (Plan)
            FOUR = ["ME DUAL","ME SIM","SE SIM","MCC SIM"]
            piv_four = (dfp[dfp["kategori_detay"].isin(FOUR)]
                        .groupby(["ay","kategori_detay"])["sure_saat"].sum()
                        .unstack(fill_value=0.0).sort_index())
            for k in FOUR:
                if k not in piv_four.columns: piv_four[k] = 0.0
            piv_four = piv_four[FOUR]
            st.bar_chart(piv_four, height=280, use_container_width=True)
            st.dataframe(_format_time_cols(piv_four.assign(TOPLAM=piv_four.sum(axis=1)).reset_index().rename(columns={"ay":"Ay"})),
                         use_container_width=True)

    # Hızlı Excel (görseller opsiyonel)
    st.markdown("### 📥 Excel Raporu (Hızlı)")
    add_images = st.toggle("Excel'e gömülü grafik görüntüsü ekle (yavaşlatır)", False)
    _download_excel_fast(st, df_cmp, aylik, tarih1, tarih2, add_images)


# -------------------------------------------------------------------
# Hızlı Excel dışa aktarım
# -------------------------------------------------------------------
def _download_excel_fast(st, df_cmp, aylik, tarih1, tarih2, add_images: bool):
    import io
    import matplotlib.pyplot as plt
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    import tempfile, os

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Günlük
        out_daily = df_cmp.copy()
        out_daily2 = out_daily.copy()
        out_daily2["Planlanan"]   = out_daily2["Planlanan"].apply(_fmt_hhmmss)
        out_daily2["Gerçekleşen"] = out_daily2["Gerçekleşen"].apply(_fmt_hhmmss)
        out_daily2["Fark"]        = out_daily2["Fark"].apply(_fmt_hhmm_signed)
        out_daily2.to_excel(writer, index=False, sheet_name="Gunluk")

        # Aylık
        aylik2 = aylik.copy()
        aylik2["Planlanan"]   = aylik2["Planlanan"].apply(_fmt_hhmmss)
        aylik2["Gerçekleşen"] = aylik2["Gerçekleşen"].apply(_fmt_hhmmss)
        aylik2["Fark"]        = aylik2["Fark"].apply(_fmt_hhmm_signed)
        aylik2.to_excel(writer, index=False, sheet_name="Aylik")

        # Özet
        pd.DataFrame([{
            "Aralık": f"{tarih1.date()} – {tarih2.date()}",
            "Toplam Plan": _fmt_hhmmss(df_cmp["Planlanan"].sum()),
            "Toplam Gerçekleşen": _fmt_hhmmss(df_cmp["Gerçekleşen"].sum()),
            "Fark": _fmt_hhmm_signed(df_cmp["Gerçekleşen"].sum() - df_cmp["Planlanan"].sum())
        }]).to_excel(writer, index=False, sheet_name="Ozet")

    if add_images:
        # küçük görseller üret
        tmpdir = tempfile.mkdtemp()
        path_daily = os.path.join(tmpdir, "daily.png")
        path_month = os.path.join(tmpdir, "monthly.png")

        plt.figure(figsize=(8,3))
        plt.plot(df_cmp["Tarih"], df_cmp["Planlanan"], marker=".")
        plt.plot(df_cmp["Tarih"], df_cmp["Gerçekleşen"], marker=".")
        plt.title("Planlanan vs Gerçekleşen (Günlük)"); plt.tight_layout(); plt.savefig(path_daily); plt.close()

        plt.figure(figsize=(8,3))
        aylik.set_index("Ay")[["Planlanan","Gerçekleşen"]].plot(kind="bar", ax=plt.gca())
        plt.title("Planlanan vs Gerçekleşen (Aylık)"); plt.tight_layout(); plt.savefig(path_month); plt.close()

        # resimleri ekle
        buf.seek(0)
        wb = load_workbook(buf)
        if "Gunluk" in wb.sheetnames:
            ws = wb["Gunluk"]; ws.add_image(XLImage(path_daily), "F2")
        if "Aylik" in wb.sheetnames:
            ws = wb["Aylik"]; ws.add_image(XLImage(path_month), "F2")
        out = io.BytesIO(); wb.save(out); out.seek(0)
        st.download_button(
            "📥 Excel indir (görsellerle)",
            data=out,
            file_name=f"plan_vs_gercek_{tarih1.date()}_{tarih2.date()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_excel_img"
        )
    else:
        buf.seek(0)
        st.download_button(
            "📥 Excel indir (hızlı)",
            data=buf,
            file_name=f"plan_vs_gercek_{tarih1.date()}_{tarih2.date()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_excel_fast"
        )
