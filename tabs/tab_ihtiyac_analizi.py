import re
import pandas as pd


def tab_ihtiyac_analizi(st, conn):
    st.subheader("📈 Tarihsel Uçuş Süre Analizi")

    # --- yardımcılar ---
    def _fmt_hhmmss(hours_float) -> str:
        try:
            sec = int(round(float(hours_float) * 3600))
        except Exception:
            sec = 0
        hh, rem = divmod(sec, 3600)
        mm, ss = divmod(rem, 60)
        return f"{hh:02d}:{mm:02d}:{ss:02d}"

    def _format_time_cols(df, cols=None):
        df2 = df.copy()
        if cols is None:
            cols = [c for c in df2.columns
                    if c.lower() in ["sure_saat", "toplam saat", "toplam", "kalan_saat", "gunluk_takvim", "y"]
                    or c.startswith("A/C") or c in ["ME DUAL", "ME SIM", "SE SIM", "MCC SIM"]]
        for col in cols:
            if col in df2.columns:
                df2[col] = df2[col].apply(_fmt_hhmmss)
        return df2

    df = pd.read_sql_query(
        "SELECT plan_tarihi, sure, gorev_tipi FROM ucus_planlari",
        conn, parse_dates=["plan_tarihi"]
    )
    if df.empty:
        st.warning("Veri bulunamadı.")
        return

    # Süre → saat
    df["sure_saat"] = pd.to_timedelta(df["sure"]).dt.total_seconds() / 3600

    # === TARİH ARALIĞI ===
    min_date = df["plan_tarihi"].min()
    max_date = df["plan_tarihi"].max()
    col1, col2 = st.columns(2)
    with col1:
        tarih1 = st.date_input("Başlangıç Tarihi", min_value=min_date, max_value=max_date, value=min_date)
    with col2:
        tarih2 = st.date_input("Bitiş Tarihi", min_value=min_date, max_value=max_date, value=max_date)
    tarih1 = pd.to_datetime(tarih1); tarih2 = pd.to_datetime(tarih2)

    # Aralık filtresi
    mask = (df["plan_tarihi"] >= tarih1) & (df["plan_tarihi"] <= tarih2)
    df_aralik = df[mask].copy()
    if df_aralik.empty:
        st.warning("Seçilen aralıkta veri bulunamadı.")
        return

    # --- KATEGORİ HARİTASI ---
    def _norm(s: str) -> str:
        return re.sub(r"\s+", " ", str(s).strip().upper())

    # A/C detay anahtarları (kalsın; toplam A/C'de kullanabiliriz)
    DA20_KEYS    = {"SE DUAL DA", "SE PIC"}
    SONACA_KEYS  = {"SE DUAL SONACA"}
    ME_DUAL_KEYS = {"ME DUAL"}
    AUPRT_KEYS   = {"AUPRT"}

    # İSTEDİĞİN 4 GÖREVİ bire bir isimleriyle üret
    def map_kategori(gorev_tipi: str) -> str:
        s = _norm(gorev_tipi)
        if s == "ME DUAL": return "ME DUAL"
        if s.startswith("ME SIM"):  return "ME SIM"
        if s.startswith("SE SIM"):  return "SE SIM"
        if s.startswith("MCC SIM"): return "MCC SIM"
        # Diğer A/C kırılımları (opsiyonel, görünmesi faydalı olabilir)
        if s in DA20_KEYS:    return "A/C – DA20"
        if s in SONACA_KEYS:  return "A/C – SONACA"
        if s in AUPRT_KEYS:   return "A/C – AUPRT"
        return "A/C – DİĞER"

    df_aralik["kategori_detay"] = df_aralik["gorev_tipi"].apply(map_kategori)
    df_aralik["ay"] = df_aralik["plan_tarihi"].dt.to_period("M").dt.to_timestamp()

    # ---- Zekî özet ----
    gunluk_top = df_aralik.groupby("plan_tarihi")["sure_saat"].sum()
    toplam_saat = gunluk_top.sum()
    ort_gunluk  = gunluk_top.mean()
    min_gun     = gunluk_top.min()
    max_gun     = gunluk_top.max()
    hic_ucus_olmayan = pd.date_range(tarih1, tarih2).difference(df_aralik["plan_tarihi"].dt.normalize().unique())
    populer_tip = df_aralik.groupby("gorev_tipi")["sure_saat"].sum().idxmax() if not df_aralik.empty else None

    st.success(
        f"**Tarih Aralığı:** {tarih1.date()} – {tarih2.date()}  \n"
        f"- Toplam Uçuş Süresi: **{_fmt_hhmmss(toplam_saat)}**  \n"
        f"- Günlük Ortalama: **{_fmt_hhmmss(ort_gunluk)}**  \n"
        f"- En yüksek gün: **{_fmt_hhmmss(max_gun)}**  \n"
        f"- En düşük (uçulan) gün: **{_fmt_hhmmss(min_gun)}**  \n"
        f"- En çok uçulan görev tipi: **{populer_tip}**  \n"
        f"- Uçuş olmayan günler: {'Yok' if len(hic_ucus_olmayan)==0 else ', '.join(str(g.date()) for g in hic_ucus_olmayan[:5])}"
    )

    # ==== Günlük toplam ====
    st.markdown("## 🔷 Toplam Uçuş Süresi (Günlük)")
    df_gunluk_total = gunluk_top.reset_index().sort_values("plan_tarihi")
    st.line_chart(df_gunluk_total.set_index("plan_tarihi"), height=300, use_container_width=True)
    st.dataframe(
        _format_time_cols(df_gunluk_total.rename(columns={"sure_saat":"Toplam Saat"})),
        use_container_width=True
    )

    # === A/C (Toplam) vs SIM (Toplam) — Aylık (çifte sayım yok) ===
    st.markdown("## ✈️ vs 🧪 A/C (Toplam) — SIM(Toplam) · Aylık")
    SIM_SET = {"ME SIM", "SE SIM", "MCC SIM"}
    aylik_ac  = (df_aralik[~df_aralik["kategori_detay"].isin(SIM_SET)]
                 .groupby("ay")["sure_saat"].sum())
    aylik_sim = (df_aralik[df_aralik["kategori_detay"].isin(SIM_SET)]
                 .groupby("ay")["sure_saat"].sum())
    piv_ac_sim = pd.concat([aylik_ac.rename("A/C"), aylik_sim.rename("SIM")], axis=1).fillna(0.0).sort_index()

    st.bar_chart(piv_ac_sim, height=330, use_container_width=True)
    st.dataframe(_format_time_cols(piv_ac_sim.reset_index().rename(columns={"ay":"Ay"})), use_container_width=True)

    # === A/C Detay — Aylık (opsiyonel görünüm) ===
    st.markdown("### 🗓️ A/C Detay — Aylık Toplam")
    piv_ac_detay = (df_aralik[~df_aralik["kategori_detay"].isin(SIM_SET)]
                    .groupby(["ay","kategori_detay"])["sure_saat"].sum()
                    .unstack(fill_value=0.0).sort_index())
    if not piv_ac_detay.empty:
        st.bar_chart(piv_ac_detay, height=300, use_container_width=True)
        st.dataframe(
            _format_time_cols(piv_ac_detay.assign(TOPLAM=piv_ac_detay.sum(axis=1)).reset_index().rename(columns={"ay":"Ay"})),
            use_container_width=True
        )
    else:
        st.caption("A/C detay bulunamadı.")

    # === 🧩 Seçili 4 Görev — Aylık (ME DUAL, ME SIM, SE SIM, MCC SIM) ===
    st.markdown("### 🧩 Seçili 4 Görev — Aylık (ME DUAL / ME SIM / SE SIM / MCC SIM)")
    FOUR = ["ME DUAL", "ME SIM", "SE SIM", "MCC SIM"]
    piv_four = (df_aralik[df_aralik["kategori_detay"].isin(FOUR)]
                .groupby(["ay","kategori_detay"])["sure_saat"].sum()
                .unstack(fill_value=0.0).sort_index())
    if not piv_four.empty:
        # kolon sırası sabit
        for k in FOUR:
            if k not in piv_four.columns: piv_four[k] = 0.0
        piv_four = piv_four[FOUR]
        st.bar_chart(piv_four, height=300, use_container_width=True)
        st.dataframe(
            _format_time_cols(piv_four.assign(TOPLAM=piv_four.sum(axis=1)).reset_index().rename(columns={"ay":"Ay"})),
            use_container_width=True
        )
    else:
        st.caption("Seçili dört görev için veri yok.")

    # === Görev tipleri alt grafikler ===
    gorev_tipleri = sorted(df_aralik["gorev_tipi"].dropna().unique().tolist())
    selected_tips = st.multiselect("Gösterilecek Görev Tipleri", options=gorev_tipleri, default=gorev_tipleri, key="tarihsel_gorev_tipleri")
    for tip in selected_tips:
        st.markdown(f"---\n## {tip}")
        df_tip = df_aralik[df_aralik["gorev_tipi"] == tip]
        df_gunluk = df_tip.groupby("plan_tarihi")["sure_saat"].sum().reset_index().sort_values("plan_tarihi")
        st.line_chart(df_gunluk.set_index("plan_tarihi"), height=250, use_container_width=True)
        st.dataframe(_format_time_cols(df_gunluk.rename(columns={"sure_saat":"Toplam Saat"})), use_container_width=True)
        st.info(f"- Toplam: {_fmt_hhmmss(df_tip['sure_saat'].sum())} | "
                f"Ortalama: {_fmt_hhmmss(df_gunluk['sure_saat'].mean())} | "
                f"Gün sayısı: {df_gunluk['plan_tarihi'].nunique()}")

    # ==== Excel Rapor ====
    st.markdown("### 📥 Excel Raporu (Grafik+Özet) İndir")
    excel_rapor_grafikli_indir(st, df_aralik, df_gunluk_total, selected_tips, tarih1, tarih2)

    # Akıllı özet panel
    tarihsel_akilli_ozet_panel(st, conn, tarih1, tarih2, df_aralik)



def excel_rapor_grafikli_indir(st, df_aralik, df_gunluk_total, selected_tips, tarih1, tarih2):
    import matplotlib.pyplot as plt
    import tempfile, os, io, re
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage

    # --- yardımcılar ---
    def _fmt_hhmmss(hours_float) -> str:
        try:
            sec = int(round(float(hours_float) * 3600))
        except Exception:
            sec = 0
        hh, rem = divmod(sec, 3600)
        mm, ss = divmod(rem, 60)
        return f"{hh:02d}:{mm:02d}:{ss:02d}"

    def _format_time_cols(df, cols=None):
        df2 = df.copy()
        if cols is None:
            cols = [c for c in df2.columns
                    if c.lower() in ["sure_saat", "toplam saat", "toplam", "kalan_saat", "gunluk_takvim", "y"]
                    or c.startswith("A/C") or c in ["ME DUAL", "ME SIM", "SE SIM", "MCC SIM"]]
        for col in cols:
            if col in df2.columns:
                df2[col] = df2[col].apply(_fmt_hhmmss)
        return df2

    tmpdir = tempfile.mkdtemp()

    # --- kategori haritası (4 görev + diğerleri) ---
    def _norm(s: str) -> str:
        return re.sub(r"\s+", " ", str(s).strip().upper())

    DA20_KEYS    = {"SE DUAL DA", "SE PIC"}
    SONACA_KEYS  = {"SE DUAL SONACA"}
    AUPRT_KEYS   = {"AUPRT"}

    def map_kategori(gorev_tipi: str) -> str:
        s = _norm(gorev_tipi)
        if s == "ME DUAL": return "ME DUAL"
        if s.startswith("ME SIM"):  return "ME SIM"
        if s.startswith("SE SIM"):  return "SE SIM"
        if s.startswith("MCC SIM"): return "MCC SIM"
        if s in DA20_KEYS:    return "A/C – DA20"
        if s in SONACA_KEYS:  return "A/C – SONACA"
        if s in AUPRT_KEYS:   return "A/C – AUPRT"
        return "A/C – DİĞER"

    dfa = df_aralik.copy()
    dfa["kategori_detay"] = dfa["gorev_tipi"].apply(map_kategori)
    dfa["ay"] = dfa["plan_tarihi"].dt.to_period("M").dt.to_timestamp()

    # --- görselleri üret ---
    path_total = os.path.join(tmpdir, "tarihsel_total.png")
    plt.figure(figsize=(8,3))
    plt.plot(df_gunluk_total["plan_tarihi"], df_gunluk_total["sure_saat"], marker="o")
    plt.title("Toplam Uçuş Süresi (Günlük)")
    plt.xlabel("Tarih"); plt.ylabel("Saat"); plt.grid(True, linestyle="--", alpha=0.5)
    plt.tight_layout(); plt.savefig(path_total); plt.close()

    # A/C (toplam) vs SIM (toplam)
    SIM_SET = {"ME SIM", "SE SIM", "MCC SIM"}
    aylik_ac  = (dfa[~dfa["kategori_detay"].isin(SIM_SET)].groupby("ay")["sure_saat"].sum())
    aylik_sim = (dfa[dfa["kategori_detay"].isin(SIM_SET)].groupby("ay")["sure_saat"].sum())
    piv_ac_sim  = pd.concat([aylik_ac.rename("A/C"), aylik_sim.rename("SIM")], axis=1).fillna(0.0).sort_index()

    path_monthly_cat = os.path.join(tmpdir, "aylik_ac_vs_sim.png")
    plt.figure(figsize=(9,3.2))
    piv_ac_sim.plot(kind="bar", ax=plt.gca())
    plt.title("A/C (Toplam) vs SIM (Toplam) — Aylık")
    plt.xlabel("Ay"); plt.ylabel("Saat"); plt.tight_layout()
    plt.savefig(path_monthly_cat); plt.close()

    # A/C detay aylık
    piv_ac_detay = (dfa[~dfa["kategori_detay"].isin(SIM_SET)]
                    .groupby(["ay","kategori_detay"])["sure_saat"]
                    .sum().unstack(fill_value=0.0).sort_index())
    path_monthly_ac_detail = os.path.join(tmpdir, "aylik_ac_detay.png")
    if not piv_ac_detay.empty:
        plt.figure(figsize=(9,3.2))
        piv_ac_detay.plot(kind="bar", ax=plt.gca())
        plt.title("A/C Detay — Aylık")
        plt.xlabel("Ay"); plt.ylabel("Saat"); plt.tight_layout()
        plt.savefig(path_monthly_ac_detail); plt.close()

    # SIM detay aylık (3 başlık)
    piv_sim_detay = (dfa[dfa["kategori_detay"].isin(SIM_SET)]
                     .groupby(["ay","kategori_detay"])["sure_saat"]
                     .sum().unstack(fill_value=0.0).sort_index())
    path_monthly_sim_detail = os.path.join(tmpdir, "aylik_sim_detay.png")
    if not piv_sim_detay.empty:
        plt.figure(figsize=(9,3.2))
        piv_sim_detay.plot(kind="bar", ax=plt.gca())
        plt.title("SIM Detay (ME SIM / SE SIM / MCC SIM) — Aylık")
        plt.xlabel("Ay"); plt.ylabel("Saat"); plt.tight_layout()
        plt.savefig(path_monthly_sim_detail); plt.close()

    # Seçili 4 görev aylık (ME DUAL + 3 SIM)
    FOUR = ["ME DUAL", "ME SIM", "SE SIM", "MCC SIM"]
    piv_four = (dfa[dfa["kategori_detay"].isin(FOUR)]
                .groupby(["ay","kategori_detay"])["sure_saat"]
                .sum().unstack(fill_value=0.0).sort_index())
    path_monthly_four = os.path.join(tmpdir, "aylik_4_gorev.png")
    if not piv_four.empty:
        for k in FOUR:
            if k not in piv_four.columns: piv_four[k] = 0.0
        piv_four = piv_four[FOUR]
        plt.figure(figsize=(9,3.2))
        piv_four.plot(kind="bar", ax=plt.gca())
        plt.title("Seçili 4 Görev — Aylık (ME DUAL / ME SIM / SE SIM / MCC SIM)")
        plt.xlabel("Ay"); plt.ylabel("Saat"); plt.tight_layout()
        plt.savefig(path_monthly_four); plt.close()

    # --- EXCEL ---
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        # Özet (saat alanları HH:MM:SS metin)
        summary = {
            "Tarih Başlangıç": str(tarih1.date()),
            "Tarih Bitiş": str(tarih2.date()),
            "Toplam Uçuş Saati": _fmt_hhmmss(dfa["sure_saat"].sum()),
            "Günlük Ortalama": _fmt_hhmmss(dfa.groupby("plan_tarihi")["sure_saat"].sum().mean()),
            "En çok uçulan görev tipi": (dfa.groupby("gorev_tipi")["sure_saat"].sum().idxmax() if not dfa.empty else "-"),
        }
        pd.DataFrame([summary]).to_excel(writer, index=False, sheet_name="Özet")

        # Günlük toplam (HH:MM:SS)
        df_gt = df_gunluk_total.rename(columns={"sure_saat":"Toplam Saat"}).copy()
        df_gt["Toplam Saat"] = df_gt["Toplam Saat"].apply(_fmt_hhmmss)
        df_gt.to_excel(writer, index=False, sheet_name="Günlük Toplam")

        # A/C vs SIM (HH:MM:SS)
        acsim_x = piv_ac_sim.reset_index().rename(columns={"ay":"Ay"}).copy()
        for col in [col for col in acsim_x.columns if col != "Ay"]:
            acsim_x[col] = acsim_x[col].apply(_fmt_hhmmss)
        acsim_x.to_excel(writer, index=False, sheet_name="Aylık AC_vs_SIM")

        # A/C detay (HH:MM:SS)
        if not piv_ac_detay.empty:
            acd_x = piv_ac_detay.reset_index().rename(columns={"ay":"Ay"}).copy()
            for col in [col for col in acd_x.columns if col != "Ay"]:
                acd_x[col] = acd_x[col].apply(_fmt_hhmmss)
            acd_x.to_excel(writer, index=False, sheet_name="Aylık AC Detay")

        # SIM detay (HH:MM:SS)
        if not piv_sim_detay.empty:
            simd_x = piv_sim_detay.reset_index().rename(columns={"ay":"Ay"}).copy()
            for col in [col for col in simd_x.columns if col != "Ay"]:
                simd_x[col] = simd_x[col].apply(_fmt_hhmmss)
            simd_x.to_excel(writer, index=False, sheet_name="Aylık SIM Detay")

        # 4 Görev (HH:MM:SS)
        if not piv_four.empty:
            four_x = piv_four.reset_index().rename(columns={"ay":"Ay"}).copy()
            # kolon sırasını sabitle
            for k in ["ME DUAL", "ME SIM", "SE SIM", "MCC SIM"]:
                if k not in four_x.columns: four_x[k] = 0.0
            four_x = four_x[["Ay","ME DUAL","ME SIM","SE SIM","MCC SIM"]]
            for col in [col for col in four_x.columns if col != "Ay"]:
                four_x[col] = four_x[col].apply(_fmt_hhmmss)
            four_x.to_excel(writer, index=False, sheet_name="Aylık 4 Görev")

        # Her görev tipi (HH:MM:SS)
        for tip in selected_tips:
            dft = dfa[dfa["gorev_tipi"] == tip]
            dfg = dft.groupby("plan_tarihi")["sure_saat"].sum().reset_index()
            dfg = dfg.rename(columns={"sure_saat":"Toplam Saat"})
            dfg["Toplam Saat"] = dfg["Toplam Saat"].apply(_fmt_hhmmss)
            dfg.to_excel(writer, index=False, sheet_name=f"Tip_{tip[:25]}")

    # görselleri ekle
    excel_buffer.seek(0)
    wb = load_workbook(excel_buffer)
    wb["Günlük Toplam"].add_image(XLImage(path_total), "E2")
    wb["Aylık AC_vs_SIM"].add_image(XLImage(path_monthly_cat), "E2")
    if "Aylık AC Detay" in wb.sheetnames and os.path.exists(path_monthly_ac_detail):
        wb["Aylık AC Detay"].add_image(XLImage(path_monthly_ac_detail), "E2")
    if "Aylık SIM Detay" in wb.sheetnames and os.path.exists(path_monthly_sim_detail):
        wb["Aylık SIM Detay"].add_image(XLImage(path_monthly_sim_detail), "E2")
    if "Aylık 4 Görev" in wb.sheetnames and os.path.exists(path_monthly_four):
        wb["Aylık 4 Görev"].add_image(XLImage(path_monthly_four), "E2")

    out_buf = io.BytesIO()
    wb.save(out_buf); out_buf.seek(0)
    st.download_button(
        label="📥 Seçilen Tarih Aralığına Göre Grafikli Excel Raporu İndir",
        data=out_buf,
        file_name=f"ucus_analiz_{tarih1.date()}_{tarih2.date()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )




def tarihsel_akilli_ozet_panel(st, conn, tarih1, tarih2, df_aralik):
    import pandas as pd, re, io
    from datetime import date

    def _to_hours(s):
        try:
            if pd.isna(s) or str(s).strip() in ("", "nan"): return 0.0
            return pd.to_timedelta(str(s)).total_seconds()/3600.0
        except Exception:
            try: return float(s)
            except Exception: return 0.0

    def _fmt_hhmmss(h):
        try:
            sec = int(round(float(h)*3600))
        except Exception:
            sec = 0
        hh, rem = divmod(sec, 3600); mm, ss = divmod(rem, 60)
        return f"{hh:02d}:{mm:02d}:{ss:02d}"

    def _format_time_cols(df, cols=None):
        df2 = df.copy()
        if cols is None:
            cols = [c for c in df2.columns
                    if c.lower() in ["sure_saat", "toplam saat", "toplam", "kalan_saat", "gunluk_takvim", "y"]
                    or c.startswith("A/C") or c in ["ME DUAL", "ME SIM", "SE SIM", "MCC SIM"]]
        for col in cols:
            if col in df2.columns:
                df2[col] = df2[col].apply(_fmt_hhmmss)
        return df2

    if df_aralik is None or df_aralik.empty:
        st.warning("Seçilen tarih aralığında veri bulunamadı."); return

    rng_min = pd.to_datetime(tarih1).normalize(); rng_max = pd.to_datetime(tarih2).normalize()
    df_range = df_aralik.copy()
    df_range["plan_tarihi"] = pd.to_datetime(df_range["plan_tarihi"]).dt.normalize()
    if "sure_saat" not in df_range.columns and "sure" in df_range.columns:
        df_range["sure_saat"] = df_range["sure"].apply(_to_hours)

    gunluk = df_range.groupby("plan_tarihi")["sure_saat"].sum().sort_index()
    total_hours = gunluk.sum()
    flown_days  = gunluk.index.nunique()
    full_range  = pd.date_range(rng_min, rng_max, freq="D").normalize()
    not_flown   = max(len(full_range) - flown_days, 0)

    today = pd.Timestamp(date.today()).normalize()
    start_point = max(today, rng_min)
    if start_point > rng_max:
        remaining_total = 0.0; remaining_days = 0
        kalan_tip = pd.DataFrame(columns=["gorev_tipi","kalan_saat"])
    else:
        kalan_df = df_range[df_range["plan_tarihi"] >= start_point]
        remaining_total = kalan_df["sure_saat"].sum()
        remaining_days  = (rng_max - start_point).days + 1
        kalan_tip = (kalan_df.assign(gorev_tipi=lambda d: d["gorev_tipi"].fillna("Bilinmiyor"))
                              .groupby("gorev_tipi", dropna=False)["sure_saat"].sum()
                              .reset_index().sort_values("sure_saat", ascending=False)
                              .rename(columns={"sure_saat":"kalan_saat"}))
    hedef_gun_takvim = (remaining_total/remaining_days) if remaining_days>0 else 0.0
    kalan_tip["gunluk_takvim"] = kalan_tip["kalan_saat"]/max(remaining_days,1)

    st.markdown(f"### 🎯 Akıllı Özet — {rng_min.date()}–{rng_max.date()} (HH:MM:SS)")
    c1,c2,c3 = st.columns(3)
    c1.metric("Toplam Plan", _fmt_hhmmss(total_hours))
    c2.metric("Uçulan Gün", f"{flown_days} gün")
    c3.metric("Uçulmayan Gün", f"{not_flown} gün")
    st.info(f"**Kalan plan ({start_point.date()} → {rng_max.date()}): {_fmt_hhmmss(remaining_total)}**")
    st.metric("Gerekli günlük (takvim)", _fmt_hhmmss(hedef_gun_takvim))

    st.markdown("### 🧮 Görev Tipi Bazında Günlük Gerekli (Takvim) — HH:MM:SS")
    if not kalan_tip.empty:
        df_show = kalan_tip[["gorev_tipi","kalan_saat","gunluk_takvim"]].copy()
        df_show.columns = ["Görev Tipi","Kalan Saat","Günlük (Takvim)"]
        st.dataframe(_format_time_cols(df_show, cols=["Kalan Saat","Günlük (Takvim)"]), use_container_width=True)
    else:
        st.caption("Seçili aralık için kalan görev tipi hedefi yok.")

    if not gunluk.empty:
        st.caption(f"Özet — Toplam: {_fmt_hhmmss(total_hours)} | "
                   f"Günlük ort. (uçulan): {_fmt_hhmmss(gunluk.mean())} | "
                   f"En yüksek gün: {_fmt_hhmmss(gunluk.max())}")

    # --- KATEGORİ & A/C / SIM / 4 Görev — Aylık ---
    def _norm(s: str) -> str:
        return re.sub(r"\s+", " ", str(s).strip().upper())

    def map_kategori(gorev_tipi: str) -> str:
        s = _norm(gorev_tipi)
        if s == "ME DUAL": return "ME DUAL"
        if s.startswith("ME SIM"):  return "ME SIM"
        if s.startswith("SE SIM"):  return "SE SIM"
        if s.startswith("MCC SIM"): return "MCC SIM"
        if s in {"SE DUAL DA","SE PIC"}:        return "A/C – DA20"
        if s in {"SE DUAL SONACA"}:            return "A/C – SONACA"
        if s in {"AUPRT"}:                      return "A/C – AUPRT"
        return "A/C – DİĞER"

    dfr = df_range.copy()
    dfr["kategori_detay"] = dfr["gorev_tipi"].apply(map_kategori)
    dfr["ay"] = dfr["plan_tarihi"].dt.to_period("M").dt.to_timestamp()

    SIM_SET = {"ME SIM", "SE SIM", "MCC SIM"}

    # A/C vs SIM
    aylik_ac  = (dfr[~dfr["kategori_detay"].isin(SIM_SET)]
                 .groupby("ay")["sure_saat"].sum())
    aylik_sim = (dfr[dfr["kategori_detay"].isin(SIM_SET)]
                 .groupby("ay")["sure_saat"].sum())
    piv_ac_sim = pd.concat([aylik_ac.rename("A/C"), aylik_sim.rename("SIM")], axis=1).fillna(0.0).sort_index()

    st.markdown("### 📊 A/C (Toplam) vs SIM (Toplam) — Aylık")
    st.bar_chart(piv_ac_sim, height=300, use_container_width=True)
    st.dataframe(_format_time_cols(piv_ac_sim.reset_index().rename(columns={"ay":"Ay"})), use_container_width=True)

    # A/C Detay
    piv_ac_detay = (dfr[~dfr["kategori_detay"].isin(SIM_SET)]
                    .groupby(["ay","kategori_detay"])["sure_saat"].sum()
                    .unstack(fill_value=0.0).sort_index())
    if not piv_ac_detay.empty:
        st.markdown("### 📊 A/C Detay — Aylık")
        st.bar_chart(piv_ac_detay, height=300, use_container_width=True)
        st.dataframe(_format_time_cols(piv_ac_detay.assign(TOPLAM=piv_ac_detay.sum(axis=1)).reset_index().rename(columns={"ay":"Ay"})),
                     use_container_width=True)

    # SIM Detay (ME SIM / SE SIM / MCC SIM)
    piv_sim_detay = (dfr[dfr["kategori_detay"].isin(SIM_SET)]
                     .groupby(["ay","kategori_detay"])["sure_saat"].sum()
                     .unstack(fill_value=0.0).sort_index())
    if not piv_sim_detay.empty:
        st.markdown("### 🧪 SIM Detay — Aylık (ME SIM / SE SIM / MCC SIM)")
        st.bar_chart(piv_sim_detay, height=300, use_container_width=True)
        st.dataframe(_format_time_cols(piv_sim_detay.assign(TOPLAM=piv_sim_detay.sum(axis=1)).reset_index().rename(columns={"ay":"Ay"})),
                     use_container_width=True)

    # Seçili 4 Görev (ME DUAL + 3 SIM)
    FOUR = ["ME DUAL", "ME SIM", "SE SIM", "MCC SIM"]
    piv_four = (dfr[dfr["kategori_detay"].isin(FOUR)]
                .groupby(["ay","kategori_detay"])["sure_saat"].sum()
                .unstack(fill_value=0.0).sort_index())
    if not piv_four.empty:
        for k in FOUR:
            if k not in piv_four.columns: piv_four[k] = 0.0
        piv_four = piv_four[FOUR]
        st.markdown("### 🧩 Seçili 4 Görev — Aylık")
        st.bar_chart(piv_four, height=300, use_container_width=True)
        st.dataframe(_format_time_cols(piv_four.assign(TOPLAM=piv_four.sum(axis=1)).reset_index().rename(columns={"ay":"Ay"})),
                     use_container_width=True)

    # --- Excel (Akıllı Özet) — HH:MM:SS ---
    st.markdown("#### 📥 Akıllı Özet Raporu (Excel) — Seçilen Aralık")
    try:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            pd.DataFrame([
                {"Alan":"Seçili Aralık","Değer":f"{rng_min.date()} – {rng_max.date()}"},
                {"Alan":"Toplam Plan","Değer":_fmt_hhmmss(total_hours)},
                {"Alan":"Uçulan Gün","Değer":f"{flown_days}"},
                {"Alan":"Uçulmayan Gün","Değer":f"{not_flown}"},
                {"Alan":"Kalan Plan","Değer":_fmt_hhmmss(remaining_total)},
                {"Alan":"Gerekli Günlük (Takvim)","Değer":_fmt_hhmmss(hedef_gun_takvim)},
                {"Alan":"Özet (uçulan gün ort.)","Değer":_fmt_hhmmss(gunluk.mean() if len(gunluk)>0 else 0.0)},
                {"Alan":"Özet (en yüksek gün)","Değer":_fmt_hhmmss(gunluk.max() if len(gunluk)>0 else 0.0)},
            ]).to_excel(writer, index=False, sheet_name="Akilli_Ozet")

            if not piv_ac_sim.empty:
                df1 = piv_ac_sim.reset_index().rename(columns={"ay":"Ay"})
                for col in [col for col in df1.columns if col != "Ay"]:
                    df1[col] = df1[col].apply(_fmt_hhmmss)
                df1.to_excel(writer, index=False, sheet_name="Aylik_AC_vs_SIM")

            if not piv_ac_detay.empty:
                df2 = piv_ac_detay.reset_index().rename(columns={"ay":"Ay"})
                for col in [col for col in df2.columns if col != "Ay"]:
                    df2[col] = df2[col].apply(_fmt_hhmmss)
                df2.to_excel(writer, index=False, sheet_name="Aylik_AC_Detay")

            if not piv_sim_detay.empty:
                df3 = piv_sim_detay.reset_index().rename(columns={"ay":"Ay"})
                for col in [col for col in df3.columns if col != "Ay"]:
                    df3[col] = df3[col].apply(_fmt_hhmmss)
                df3.to_excel(writer, index=False, sheet_name="Aylik_SIM_Detay")

            if not piv_four.empty:
                df4 = piv_four.reset_index().rename(columns={"ay":"Ay"})
                for col in [col for col in df4.columns if col != "Ay"]:
                    df4[col] = df4[col].apply(_fmt_hhmmss)
                df4.to_excel(writer, index=False, sheet_name="Aylik_4_Gorev")

            if not gunluk.empty:
                gt = gunluk.reset_index().rename(columns={"plan_tarihi":"Tarih","sure_saat":"Toplam Saat"})
                gt["Toplam Saat"] = gt["Toplam Saat"].apply(_fmt_hhmmss)
                gt.to_excel(writer, index=False, sheet_name="Gunluk_Toplam")

        buf.seek(0)
        st.download_button(
            label="📥 Akıllı Özet Raporunu İndir (Excel)",
            data=buf,
            file_name=f"akilli_ozet_{rng_min.date()}_{rng_max.date()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="akil_ozet_rapor_indir"
        )
    except Exception as e:
        st.error(f"Rapor oluşturulurken hata: {e}")
